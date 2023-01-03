import re
from datasource import *
import openpyxl
from datetime import datetime

def createNewFile(new_name, structuredData): #This is run last after all the data has been compiled
  global wb
  wb = openpyxl.Workbook() #create new workbook to store data
  sheet = wb.worksheets[0] #create sheet inside the workbook
  for rowindex, row in enumerate(structuredData): #loop through all of the final data and paste it in the cells
    for columnindex, data in enumerate(row):
        sheet[f"{openpyxl.utils.cell.get_column_letter(columnindex+1)}{rowindex+1}"] = data
  wb.create_sheet('Data')
  wb.save(new_name+'.xlsx') #save the excel file!

channelToAnalyze = input("Which youtube channel to analyze? Please enter: ")

#finds all indexes of the occurances of this string which happens to be the string before the video's details
matches = re.finditer(r"""<yt-formatted-string id="video-title" class="style-scope ytd-rich-grid-media" aria-label=""", datasourceString) 
#setting up the final storage for the data
finalConfirmedArray = []
finalOutput = [["Title", "Video #", "Views"]]

#loops through all the found video information
for match in matches:
    #takes the next 1000 characters (which happens to be the video data), as no video data will have anywhere near 1000 characters
    contentString = datasourceString[int(match.end()):int(match.end())+1000] 
    #cleaning up the data
    before = contentString.split('>')[0]
    checkStringArray = before.lower().split(' ')
    contains = False
    #putting all the video data into the data structure so we can dissect it
    for stringA in checkStringArray:
        if stringA == channelToAnalyze.lower():
            contains = True
    if contains:
        cleanedString = before[1:]
        cleanedString = cleanedString[:len(cleanedString)-1]
        finalConfirmedArray.append(cleanedString)

#dissecting the now structured string data
for index, ian in enumerate(finalConfirmedArray):
    titleAndData = ian.split(f' by {channelToAnalyze} ')
    dataArray = []

    #TITLE
    title = titleAndData[0]
    
    #VIEW COUNT
    print(titleAndData)
    viewsPartition = titleAndData[1].split(' ')
    preViewsSplit = viewsPartition[len(viewsPartition)-2]
    views = int(''.join(preViewsSplit.split(',')))
    dataArray.append(title) #title
    dataArray.append(len(finalConfirmedArray)-1-index) #video index (video number)
    dataArray.append(views) #view count
    finalOutput.append(dataArray) #appending to final data for excel to process

#inserting data into excel!
now = datetime.now()
current_time = now.strftime("%Y-%m-%d-%H%M%S")
createNewFile(f"{channelToAnalyze}-{current_time}", finalOutput)

